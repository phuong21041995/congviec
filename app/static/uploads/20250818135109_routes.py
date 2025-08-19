import os
import random
from datetime import datetime, timedelta, date
from sqlalchemy import func, or_
import io
from collections import defaultdict
from calendar import monthrange
import bleach
import json
from bleach.css_sanitizer import CSSSanitizer

from flask import (Blueprint, current_app, flash, jsonify, redirect,
                   render_template, request, send_from_directory, url_for, send_file)
from flask_login import login_required, current_user
from sqlalchemy import func, and_
from sqlalchemy.orm import joinedload
from werkzeug.utils import secure_filename
from sqlalchemy.exc import IntegrityError
import openpyxl
from ics import Calendar, Event

from app import db
from app.models import (ActionItem, Kaizen, KeyResult, Log,
                        Objective, Project, Task, UploadedFile, User, Column, Note)

bp = Blueprint('main', __name__)

# ==============================================================================
# ROUTE API TẬP TRUNG CHO UPLOAD FILE
# ==============================================================================
# Trong app/routes.py, cập nhật hàm upload_attachment

@bp.route('/api/upload-attachment', methods=['POST'])
@login_required
def upload_attachment():
    if 'file' not in request.files:
        return jsonify({'success': False, 'message': 'No file.'}), 400
    
    file = request.files['file']
    if file.filename == '':
        return jsonify({'success': False, 'message': 'Not yet choose file'}), 400

    # Lấy các tham số từ form
    task_id = request.form.get('task_id')
    note_id = request.form.get('note_id')
    action_id = request.form.get('action_item_id')
    source = request.form.get('source', 'attachment') # Mặc định là 'attachment'

    if file:
        original_filename = secure_filename(file.filename)
        timestamp = datetime.now().strftime("%Y%m%d%H%M%S")
        saved_filename = f"{timestamp}_{original_filename}"
        file_path = os.path.join(current_app.config['UPLOAD_FOLDER'], saved_filename)
        
        try:
            file.save(file_path)
            file_size = os.path.getsize(file_path)
            _, file_ext = os.path.splitext(original_filename)

            new_file = UploadedFile(
                original_filename=original_filename,
                saved_filename=saved_filename,
                file_type=file_ext.lower(),
                file_size=file_size,
                uploader_id=current_user.id,
                upload_source=source # Lưu nguồn upload vào DB
            )
            
            if task_id: new_file.task_id = int(task_id)
            if note_id: new_file.note_id = int(note_id)
            if action_id: new_file.action_item_id = int(action_id)

            db.session.add(new_file)
            db.session.commit()
            
            return jsonify({ 'success': True, 'file': new_file.to_dict() })
        except Exception as e:
            db.session.rollback()
            return jsonify({'success': False, 'message': f'Error happend: {str(e)}'}), 500
            
    return jsonify({'success': False, 'message': 'Error dont define'}), 500

# ==============================================================================
# HÀM HỖ TRỢ & EXPORT
# ==============================================================================
def get_date_range(view_mode, date_str):
    try:
        base_date = datetime.strptime(date_str, '%Y-%m-%d').date()
    except (ValueError, TypeError):
        base_date = datetime.today().date()

    if view_mode == 'day':
        start_date = end_date = base_date
        date_display = f"Day {start_date.strftime('%d/%m/%Y')}"
    elif view_mode == 'month':
        start_date = base_date.replace(day=1)
        next_month = start_date.replace(day=28) + timedelta(days=4)
        end_date = next_month - timedelta(days=next_month.day)
        date_display = f"Month {start_date.strftime('%m/%Y')}"
    else:
        start_date = base_date - timedelta(days=base_date.weekday())
        end_date = start_date + timedelta(days=6)
        date_display = f"Week {start_date.isocalendar()[1]} ({start_date.strftime('%d/%m')} - {end_date.strftime('%d/%m/%Y')})"
    return start_date, end_date, date_display

def get_time_range_from_filter(time_filter):
    today = date.today()
    if time_filter == 'today':
        start_date = end_date = today
    elif time_filter == 'this_month':
        start_date = today.replace(day=1)
        next_month = (start_date.replace(day=28) + timedelta(days=4))
        end_date = next_month - timedelta(days=next_month.day)
    elif time_filter == 'last_7_days':
        start_date = today - timedelta(days=6)
        end_date = today
    elif time_filter == 'last_30_days':
        start_date = today - timedelta(days=29)
        end_date = today
    else:
        start_date = today - timedelta(days=today.weekday())
        end_date = start_date + timedelta(days=6)
    return start_date, end_date

@bp.route('/export')
@login_required
def export_data():
    data_type = request.args.get('data_type', 'tasks')
    file_format = request.args.get('format', 'xlsx')
    
    user_filter = request.args.get('user_filter', 'all')
    project_filter = request.args.get('project_filter', 'all')
    time_filter = request.args.get('time_filter', 'this_week')
    start_date, end_date = get_time_range_from_filter(time_filter)

    if data_type == 'tasks' and file_format == 'xlsx':
        query = Task.query.filter(Task.task_date.between(start_date, end_date))
        if user_filter != 'all':
            query = query.filter(Task.who_id == user_filter)
        tasks = query.order_by(Task.task_date, Task.hour).all()
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = "Công việc"
        headers = ["Date", "Time", "Content", "PIC", "Status", "Note"]
        sheet.append(headers)
        for task in tasks:
            sheet.append([
                task.task_date.strftime('%Y-%m-%d'),
                f"{task.hour}:00" if task.hour is not None else "All day",
                task.what,
                task.assignee.username if task.assignee else "",
                task.status,
                task.note
            ])
        mem_file = io.BytesIO()
        workbook.save(mem_file)
        mem_file.seek(0)
        return send_file(mem_file, as_attachment=True, download_name='cong_viec.xlsx', mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

    elif data_type == 'tasks' and file_format == 'ics':
        query = Task.query.filter(Task.task_date.between(start_date, end_date))
        if user_filter != 'all':
            query = query.filter(Task.who_id == user_filter)
        tasks = query.all()

        cal = Calendar()
        for task in tasks:
            event = Event()
            event.name = task.what
            if task.hour is not None:
                event.begin = datetime.combine(task.task_date, datetime.min.time()).replace(hour=task.hour)
                event.duration = timedelta(hours=1)
            else:
                event.begin = task.task_date
                event.make_all_day()
            description = f"Status: {task.status}\n"
            if task.assignee:
                description += f"PIC: {task.assignee.username}\n"
            if task.note:
                description += f"Note: {task.note}"
            event.description = description
            cal.events.add(event)
        mem_file = io.BytesIO(str(cal).encode('utf-8'))
        mem_file.seek(0)
        return send_file(mem_file, as_attachment=True, download_name='lich_cong_viec.ics', mimetype='text/calendar')

    elif data_type == 'okrs' and file_format == 'xlsx':
        query = ActionItem.query.join(KeyResult).join(Objective).filter(Objective.start_date.between(start_date, end_date))
        if user_filter != 'all':
            query = query.filter(ActionItem.assignee_id == user_filter)
        if project_filter != 'all':
            query = query.filter(Objective.project_id == project_filter)
        actions = query.all()
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = "OKRs"
        headers = ["Objective", "Key Result", "Action Item", "Người thực hiện", "Ngày hết hạn", "Trạng thái", "Báo cáo"]
        sheet.append(headers)
        for action in actions:
            sheet.append([
                action.key_result.objective.content,
                action.key_result.content,
                action.content,
                action.assignee.username if action.assignee else "",
                action.due_date.strftime('%Y-%m-%d') if action.due_date else "",
                action.status,
                action.report
            ])
        mem_file = io.BytesIO()
        workbook.save(mem_file)
        mem_file.seek(0)
        return send_file(mem_file, as_attachment=True, download_name='okrs.xlsx', mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

    elif data_type == 'files' and file_format == 'xlsx':
        files = UploadedFile.query.order_by(UploadedFile.upload_date.desc()).all()
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = "Files"
        headers = ["Tên File Gốc", "Loại File", "Dung lượng (KB)", "Người tải lên", "Ngày tải lên"]
        sheet.append(headers)
        for f in files:
            local_dt = (f.upload_date + timedelta(hours=7)) if f.upload_date else None
            sheet.append([
                f.original_filename,
                f.file_type,
                (f.file_size or 0) / 1024,
                f.uploader.username if f.uploader else "",
                local_dt.strftime('%Y-%m-%d %H:%M') if local_dt else ""
            ])
        mem_file = io.BytesIO()
        workbook.save(mem_file)
        mem_file.seek(0)
        return send_file(mem_file, as_attachment=True, download_name='danh_sach_files.xlsx', mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

    flash('Loại export không hợp lệ.', 'danger')
    return redirect(request.referrer or url_for('main.index'))

# ==============================================================================
# LỊCH CÔNG VIỆC (CALENDAR)
# ==============================================================================
@bp.route('/favicon.ico')
def favicon():
    return '', 204
# Trong file app/routes.py

@bp.route('/')
@bp.route('/calendar/<string:view_mode>/<string:date_str>')
@login_required
def index(view_mode='week', date_str=None):
    # Lấy user_id từ URL, mặc định là 'all' để hiển thị tất cả
    selected_user_id = request.args.get('user_id', 'all')

    if date_str is None:
        date_str = datetime.today().strftime('%Y-%m-%d')

    # Tính toán các khoảng thời gian dựa trên view_mode
    start_date, end_date, date_display = get_date_range(view_mode, date_str)
    base_date = datetime.strptime(date_str, '%Y-%m-%d').date()

    if view_mode == 'day':
        prev_period_date = (base_date - timedelta(days=1)).strftime('%Y-%m-%d')
        next_period_date = (base_date + timedelta(days=1)).strftime('%Y-%m-%d')
    elif view_mode == 'week':
        prev_period_date = (start_date - timedelta(days=7)).strftime('%Y-%m-%d')
        next_period_date = (start_date + timedelta(days=7)).strftime('%Y-%m-%d')
    else:  # month
        prev_period_date = (start_date - timedelta(days=1)).replace(day=1).strftime('%Y-%m-%d')
        next_period_date = (end_date + timedelta(days=1)).strftime('%Y-%m-%d')

    # Chuẩn bị dữ liệu cho các dropdown điều hướng
    year = base_date.year
    first_day_of_year = datetime(year, 1, 1)
    weeks_in_year = [{'num': (first_day_of_year + timedelta(days=i*7)).isocalendar()[1], 'date_str': (first_day_of_year + timedelta(days=i*7)).strftime('%Y-%m-%d')} for i in range(53) if (first_day_of_year + timedelta(days=i*7)).year == year]
    months_in_year = [{'name': datetime(year, i, 1).strftime('%B'), 'date_str': datetime(year, i, 1).strftime('%Y-%m-%d')} for i in range(1, 13)]
    
    # Lấy dữ liệu chung
    all_users = User.query.order_by(User.username).all()
    logs = Log.query.order_by(Log.timestamp.desc()).limit(20).all()

    # Xây dựng context ban đầu sẽ được gửi tới template
    context = {
        'page_name': 'calendar',
        'view_mode': view_mode,
        'date_str': date_str,
        'date_display': date_display,
        'prev_period_date': prev_period_date,
        'next_period_date': next_period_date,
        'today_date_str': datetime.today().strftime('%Y-%m-%d'),
        'weeks_in_year': weeks_in_year,
        'months_in_year': months_in_year,
        'users': all_users,
        'logs': logs,
        'selected_user_id': selected_user_id
    }

    # Xây dựng câu truy vấn Task cơ bản
    base_query = Task.query.options(joinedload(Task.attachments), joinedload(Task.assignee))
    
    # Áp dụng bộ lọc user nếu người dùng đã chọn
    if selected_user_id != 'all':
        base_query = base_query.filter(Task.who_id == selected_user_id)

    # 1. Xác định khoảng thời gian truy vấn dữ liệu
    if view_mode == 'month':
        start_of_grid = start_date - timedelta(days=start_date.weekday())
        end_of_grid = start_of_grid + timedelta(days=41) # 6 tuần * 7 ngày
        task_query_range = base_query.filter(Task.task_date.between(start_of_grid, end_of_grid))
    else:
        task_query_range = base_query.filter(Task.task_date.between(start_date, end_date))

    # 2. Lấy tất cả task trong khoảng thời gian đã xác định
    all_tasks_in_range = task_query_range.order_by(Task.task_date.desc()).all()

    # 3. Tính toán dữ liệu SUMMARY (bao gồm cả đếm status cho biểu đồ)
    summary_data = []
    tasks_by_user = defaultdict(list)
    tasks_for_summary = [t for t in all_tasks_in_range if start_date <= t.task_date <= end_date]
    
    for task in tasks_for_summary:
        if task.assignee:
            tasks_by_user[task.assignee].append(task)
    
    for user, tasks in tasks_by_user.items():
        status_counts = defaultdict(int)
        for task in tasks:
            status_counts[task.status] += 1
        
        summary_data.append({
            'user': user.to_dict(),
            'task_count': len(tasks),
            'tasks': [t.to_dict() for t in sorted(tasks, key=lambda x: x.task_date)],
            'status_counts': dict(status_counts)
        })
   
    summary_data.sort(key=lambda x: x['user']['username'])  
    context['summary_data'] = summary_data

    # 4. Tính toán dữ liệu LỊCH (tasks_by_date)
    tasks_by_date = defaultdict(list)
    for task in all_tasks_in_range:
        tasks_by_date[task.task_date.strftime('%Y-%m-%d')].append(task.to_dict())
    context['tasks_by_date'] = tasks_by_date

    # 5. Cập nhật context riêng cho từng view mode
    if view_mode == 'month':
        context.update({
            'calendar_dates': [start_of_grid + timedelta(days=i) for i in range(42)],
            'current_month': start_date.month
        })
    else:
        context.update({'hours': range(8, 20)})
        if view_mode == 'day':
            context['week_dates'] = [start_date]
            day_names = ["Mon", "Tue", "Wed", "Thu", "Fri", "Sat", "Sun"]
            context['week_days_display'] = [(day_names[start_date.weekday()], start_date.strftime('%a'))]
            day_tasks = tasks_by_date.get(start_date.strftime('%Y-%m-%d'), [])
            context['day_stats'] = {
                status: sum(1 for t in day_tasks if t['status'] == status) 
                for status in ['Pending', 'In Progress', 'Review', 'Done', 'Drop']
            }
        else:  # 'week' view
            context['week_dates'] = [start_date + timedelta(days=i) for i in range(7)]
            context['week_days_display'] = [("Mon", "Mon"), ("Tue", "Tue"), ("Wed", "Wed"), ("Thu", "Thu"), ("Fri", "Fri"), ("Sat", "Sat"), ("Sun", "Sun")]

    return render_template('index.html', **context)
@bp.route('/save-task', methods=['POST'])
@login_required
def save_task():
    data = request.form
    task_id = data.get('taskId')
    task_date_str = data.get('taskDate')
    what = bleach.clean(data.get('taskWhat', ''))
    note = bleach.clean(data.get('taskNote', ''))
    hour_str = data.get('taskHour')
    recurrence = data.get('taskRecurrence', 'none')
    end_date_str = data.get('taskRecurrenceEndDate')
    who_id = data.get('taskWho')

    if not what or not task_date_str:
        return jsonify({'success': False, 'message': 'Miss important infor'}), 400

    try:
        task_date = datetime.strptime(task_date_str, '%Y-%m-%d').date()
        hour = int(hour_str) if hour_str else None
        recurrence_end_date = datetime.strptime(end_date_str, '%Y-%m-%d').date() if end_date_str and recurrence != 'none' else None
    except (ValueError, TypeError):
        return jsonify({'success': False, 'message': 'Formate date/time not true'}), 400


    if task_id:
        task = Task.query.get_or_404(task_id)
        task.what, task.hour, task.who_id, task.status, task.note, task.recurrence, task.recurrence_end_date = \
            what, hour, who_id or None, data['taskStatus'], note, recurrence, recurrence_end_date
        log_content = f"Updated event ID {task.id}"
    else:
        task = Task(task_date=task_date, hour=hour, what=what, who_id=who_id or None, status=data['taskStatus'], note=note, recurrence=recurrence, recurrence_end_date=recurrence_end_date)
        db.session.add(task)
        db.session.flush()
        log_content = f"Created new event"
    
    db.session.commit()
    
    db.session.add(Log(action=f"{log_content}: '{what}' at {task_date.strftime('%d/%m')}", user_id=current_user.id))
    db.session.commit()
    
    if not task_id and recurrence != 'none' and recurrence_end_date:
        current_date, original_day = task.task_date, task.task_date.day
        while True:
            if recurrence == 'daily': current_date += timedelta(days=1)
            elif recurrence == 'weekly': current_date += timedelta(days=7)
            elif recurrence == 'monthly':
                next_month, next_year = (current_date.month + 1, current_date.year)
                if next_month > 12: next_month, next_year = 1, next_year + 1
                last_day = monthrange(next_year, next_month)[1]
                current_date = date(next_year, next_month, min(original_day, last_day))

            if current_date > recurrence_end_date: break
            if not Task.query.filter_by(task_date=current_date, hour=task.hour).first():
                db.session.add(Task(task_date=current_date, hour=task.hour, what=task.what, who_id=task.who_id, status=task.status, note=task.note, recurrence='none'))
    
    db.session.commit()
    db.session.refresh(task)
    return jsonify({'success': True, 'task': task.to_dict(), 'message': 'Saved success!'})

@bp.route('/update-task-time', methods=['POST'])
@login_required
def update_task_time():
    data = request.json
    task = Task.query.get_or_404(data.get('taskId'))
    new_date = datetime.strptime(data.get('newDate'), '%Y-%m-%d').date()
    new_hour = data.get('newHour')
    
        
    task.task_date, task.hour = new_date, new_hour
    db.session.add(Log(action=f"Dời lịch CV ID {task.id}: '{task.what}' sang {new_date.strftime('%d/%m/%Y')}", user_id=current_user.id))
    db.session.commit()
    return jsonify({'success': True, 'task': task.to_dict(), 'message': 'Updated success'})

@bp.route('/delete-task/<int:task_id>', methods=['POST'])
@login_required
def delete_task(task_id):
    task = Task.query.get_or_404(task_id)
    db.session.add(Log(action=f"Deleted event ID {task.id}: '{task.what}'", user_id=current_user.id))
    db.session.delete(task)
    db.session.commit()
    return jsonify({'success': True, 'message': 'Deleted event'})

@bp.route('/uploads/<filename>')
@login_required
def uploaded_file(filename):
    return send_from_directory(current_app.config['UPLOAD_FOLDER'], filename)

@bp.route('/delete-uploaded-file/<int:file_id>', methods=['POST'])
@login_required
def delete_uploaded_file(file_id):
    uploaded_file = UploadedFile.query.get_or_404(file_id)
    try:
        file_path = os.path.join(current_app.config['UPLOAD_FOLDER'], uploaded_file.saved_filename)
        if os.path.exists(file_path):
            os.remove(file_path)
        
        db.session.delete(uploaded_file)
        db.session.commit()
        return jsonify({'success': True, 'message': 'File deleted!'})
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'message': f'Error when delete file: {str(e)}'}), 500

# ==============================================================================
# OKR & KAIZEN
# ==============================================================================
def recalculate_kr_progress(kr_id):
    kr = KeyResult.query.get(kr_id)
    if kr:
        action_items_list = list(kr.action_items)
        if action_items_list:
            total_actions = len(action_items_list)
            done_actions = sum(1 for action in action_items_list if action.status == 'Done')
            kr.current = float(done_actions)
            kr.target = float(total_actions) if total_actions > 0 else 1.0
        else:
            kr.current = 0
            kr.target = 0
        db.session.commit()
    return kr

@bp.route('/okr')
@bp.route('/okr/<string:view_mode>/<string:date_str>')
@login_required
def okr_page(view_mode='week', date_str=None):
    if date_str is None: date_str = datetime.today().strftime('%Y-%m-%d')
    start_date, end_date, date_display = get_date_range(view_mode, date_str)
    base_date = datetime.strptime(date_str, '%Y-%m-%d').date()

    if view_mode == 'week': prev_period_date, next_period_date = (start_date - timedelta(days=7)).strftime('%Y-%m-%d'), (start_date + timedelta(days=7)).strftime('%Y-%m-%d')
    elif view_mode == 'month': prev_period_date, next_period_date = (start_date - timedelta(days=1)).replace(day=1).strftime('%Y-%m-%d'), (end_date + timedelta(days=1)).strftime('%Y-%m-%d')
    else: prev_period_date, next_period_date = start_date.replace(year=start_date.year - 1).strftime('%Y-%m-%d'), start_date.replace(year=start_date.year + 1).strftime('%Y-%m-%d')
    
    objectives = Objective.query.options(joinedload(Objective.key_results).joinedload(KeyResult.action_items).joinedload(ActionItem.assignee)).filter(Objective.start_date.between(start_date, end_date)).order_by(Objective.id).all()
    kaizen_items = Kaizen.query.filter(Kaizen.week_start_date.between(start_date, end_date)).all()
    users = User.query.all()
    projects = Project.query.all()
    logs = Log.query.order_by(Log.timestamp.desc()).limit(20).all()
    
    year = base_date.year
    first_day_of_year = datetime(year, 1, 1)
    weeks_in_year = [{'num': (first_day_of_year + timedelta(days=i*7)).isocalendar()[1], 'date_str': (first_day_of_year + timedelta(days=i*7)).strftime('%Y-%m-%d')} for i in range(53) if (first_day_of_year + timedelta(days=i*7)).year == year]
    months_in_year = [{'name': f"Month {i}", 'date_str': datetime(year, i, 1).strftime('%Y-%m-%d')} for i in range(1, 13)]
    
    context = {
        'view_mode': view_mode, 'date_str': date_str, 'date_display': date_display,
        'objectives': objectives, 'kaizen_items': kaizen_items,
        'prev_period_date': prev_period_date, 'next_period_date': next_period_date,
        'weeks_in_year': weeks_in_year, 'months_in_year': months_in_year,
        'today_date_str': datetime.today().strftime('%Y-%m-%d'),
        'today_date_obj': date.today(),
        'users': users, 'projects': projects, 'page_name': 'okr',
        'logs': logs
    }
    return render_template('okr.html', **context)

@bp.route('/add-objective', methods=['POST'])
@login_required
def add_objective():
    data = request.form
    content = data.get('content')
    if not content:
        flash('Need to fill purpose.', 'danger')
        return redirect(url_for('main.okr_page', view_mode=data.get('view_mode', 'week'), date_str=data.get('start_date')))

    start_date, _, _ = get_date_range(data['view_mode'], data['start_date'])
    colors = ['#0d6efd', '#1cc88a', '#36b9cc', '#f6c23e', '#e74a3b', '#6f42c1']
    new_obj = Objective(content=content, start_date=start_date, color=random.choice(colors), owner_id=data.get('owner_id') or None, project_id=data.get('project_id') or None)
    db.session.add(new_obj)
    db.session.commit()
    
    db.session.add(Log(action=f"Create new Objective: '{new_obj.content}'", user_id=current_user.id))
    db.session.commit()
    
    flash('Created new objective!', 'success')
    return redirect(url_for('main.okr_page', view_mode=data['view_mode'], date_str=data['start_date']))

@bp.route('/add-key-result', methods=['POST'])
@login_required
def add_key_result():
    data = request.get_json()
    if not data or not data.get('objective_id') or not data.get('content'):
        return jsonify({'success': False, 'message': 'Miss information'}), 400
    
    new_kr = KeyResult(objective_id=data['objective_id'], content=data['content'], target=0, current=0)
    db.session.add(new_kr)
    db.session.commit()
    db.session.add(Log(action=f"Add KR to Objective ID {data['objective_id']}: '{data['content']}'", user_id=current_user.id))
    db.session.commit()
    
    return jsonify({ 'success': True, 'kr': { 'id': new_kr.id, 'content': new_kr.content, 'progress': 0, 'current': 0, 'target': 0, 'objective_id': new_kr.objective_id } })
    
@bp.route('/add-action-item', methods=['POST'])
@login_required
def add_action_item():
    data = request.form
    kr_id = data.get('key_result_id')
    content = data.get('content')
    if not content or not kr_id:
        return jsonify({'success': False, 'message': 'Content and KR ID needed'}), 400

    kr = KeyResult.query.get_or_404(kr_id)
    due_date = datetime.strptime(data.get('due_date'), '%Y-%m-%d').date() if data.get('due_date') else None

    new_action = ActionItem(content=content, key_result_id=kr.id, assignee_id=data.get('assignee_id') or None, due_date=due_date)
    db.session.add(new_action)
    db.session.commit()
    
    recalculate_kr_progress(kr.id)

    db.session.add(Log(action=f"Thêm Action Item vào KR ID {kr.id}: '{content}'", user_id=current_user.id))
    db.session.commit()

    return jsonify({
        'success': True, 'message': 'Added new action',
        'action': new_action.to_dict(),
        'kr_id': kr.id, 'objective_id': kr.objective.id,
        'kr_progress': kr.progress, 'obj_progress': kr.objective.progress,
        'kr_current': kr.current, 'kr_target': kr.target
    })
    
@bp.route('/update-action-status/<int:action_id>', methods=['POST'])
@login_required
def update_action_status(action_id):
    action = ActionItem.query.get_or_404(action_id)
    action.status = 'Done' if request.json.get('checked') else 'To Do'
    db.session.commit()
    
    status_text = "Hoàn thành" if action.status == 'Done' else "Chuyển về To Do"
    db.session.add(Log(action=f"{status_text} Action ID {action.id}: '{action.content}'", user_id=current_user.id))
    db.session.commit()
    
    kr = recalculate_kr_progress(action.key_result_id)
    return jsonify({
        'success': True, 'kr_id': kr.id, 'objective_id': kr.objective_id,
        'kr_progress': kr.progress, 'obj_progress': kr.objective.progress,
        'kr_current': kr.current, 'kr_target': kr.target
    })
    
@bp.route('/update-action/<int:action_id>', methods=['POST'])
@login_required
def update_action(action_id):
    action = ActionItem.query.get_or_404(action_id)
    data = request.json
    
    if data.get('content', '').strip(): action.content = data['content'].strip()
    
    due_date_str = data.get('due_date')
    if due_date_str: action.due_date = datetime.strptime(due_date_str, '%Y-%m-%d').date()
    elif due_date_str is not None: action.due_date = None

    assignee_id = data.get('assignee_id')
    if assignee_id: action.assignee_id = int(assignee_id) if assignee_id.isdigit() else None
    elif assignee_id is not None: action.assignee_id = None

    db.session.commit()
    db.session.add(Log(action=f"Updated Action ID {action.id}: '{action.content}'", user_id=current_user.id))
    db.session.commit()
    return jsonify({'success': True, 'message': 'Updated success!', 'action': action.to_dict()})

@bp.route('/save-kaizen', methods=['POST'])
@login_required
def save_kaizen():
    data = request.form
    problem = data.get('problem')
    if not problem:
        flash('Nội dung Vấn đề không được để trống.', 'danger')
        return redirect(url_for('main.okr_page', view_mode=data.get('view_mode', 'week'), date_str=data.get('date_str')))

    start_date, _, _ = get_date_range(data['view_mode'], data['date_str'])
    new_kaizen = Kaizen(week_start_date=start_date, problem=problem, solution=data['solution'])
    db.session.add(new_kaizen)
    db.session.commit()
    
    db.session.add(Log(action=f"Save new Kaizne: '{problem}'", user_id=current_user.id))
    db.session.commit()
    
    flash('Kaizend saved!', 'success')
    return redirect(url_for('main.okr_page', view_mode=data.get('view_mode', 'week'), date_str=data.get('date_str')))

@bp.route('/update/<item_type>/<int:item_id>', methods=['POST'])
@login_required
def update_item(item_type, item_id):
    model_map = {'objective': Objective, 'key_result': KeyResult, 'action_item': ActionItem}
    Model = model_map.get(item_type)
    if not Model: return jsonify({'success': False, 'message': 'Type not suitable'}), 400
    item = Model.query.get_or_404(item_id)
    new_content = request.json.get('content', '').strip()
    if new_content:
        item.content = new_content
        db.session.commit()
        db.session.add(Log(action=f"Cập nhật nội dung {item_type} ID {item.id} thành '{item.content}'", user_id=current_user.id))
        db.session.commit()
        return jsonify({'success': True, 'new_content': item.content})
    return jsonify({'success': False, 'message': 'Content cannot empty'}), 400

@bp.route('/delete/<item_type>/<int:item_id>', methods=['POST'])
@login_required
def delete_item(item_type, item_id):
    model_map = {'objective': Objective, 'key_result': KeyResult, 'action_item': ActionItem}
    Model = model_map.get(item_type)
    if not Model: return jsonify({'success': False, 'message': 'Type not suitable'}), 400
    
    item = Model.query.get_or_404(item_id)
    response_data = {'success': True}
    
    db.session.add(Log(action=f"Xóa {item_type} ID {item.id}: '{item.content}'", user_id=current_user.id))

    if item_type == 'action_item':
        kr_id = item.key_result_id
        db.session.delete(item)
        db.session.commit()
        kr = recalculate_kr_progress(kr_id)
        if kr:
            response_data.update({
                'kr_id': kr.id, 'objective_id': kr.objective_id,
                'kr_progress': kr.progress, 'obj_progress': kr.objective.progress,
                'kr_current': kr.current, 'kr_target': kr.target
            })
    elif item_type == 'key_result':
        obj = item.objective
        db.session.delete(item)
        db.session.commit()
        response_data.update({'objective_id': obj.id, 'obj_progress': obj.progress})
    else:
        db.session.delete(item)
        db.session.commit()

    return jsonify(response_data)
    
@bp.route('/action/<int:action_id>')
@login_required
def action_detail(action_id):
    action = ActionItem.query.options(joinedload(ActionItem.key_result).joinedload(KeyResult.objective)).get_or_404(action_id)
    return render_template('action_detail.html', action=action, page_name='okr')
    
# Dán code này vào file app/routes.py, thay thế cho hàm cũ

@bp.route('/action/<int:action_id>/save-report', methods=['POST'])
@login_required
def save_action_report(action_id):
    action = ActionItem.query.get_or_404(action_id)
    
    report_content = request.form.get('report_content')
    if report_content is not None:
        # SỬA LỖI: Dùng cú pháp cũ của thư viện bleach để tương thích
        allowed_tags = list(bleach.ALLOWED_TAGS) + [
            'p', 'br', 'strong', 'em', 'u', 'ol', 'ul', 'li', 
            'img', 'a', 'span', 'div'
        ]
        allowed_attrs = {
            **bleach.ALLOWED_ATTRIBUTES, 
            'img': ['src', 'alt', 'style', 'width', 'height'], 
            'a': ['href', 'title'], 
            'span':['style'], 
            'div':['style']
        }
        allowed_styles = [
            'color', 'background-color', 'text-align', 
            'font-weight', 'font-style', 'text-decoration'
        ]
        
        action.report = bleach.clean(
            report_content, 
            tags=allowed_tags, 
            attributes=allowed_attrs 
            #styles=allowed_styles
        )

    db.session.commit()
    flash('Saved!', 'success')
    return redirect(url_for('main.action_detail', action_id=action_id))

# ==============================================================================
# SEARCH, UPLOADS, PROJECTS
# ==============================================================================
@bp.route('/search')
@login_required
def search():
    query = request.args.get('q', '').strip()
    if not query:
        return render_template('search_results.html', query=query, tasks=[], objectives=[], key_results=[], actions=[])

    search_term = f"%{query}%"
    tasks = Task.query.filter(or_(Task.what.ilike(search_term), Task.note.ilike(search_term))).all()
    objectives = Objective.query.filter(Objective.content.ilike(search_term)).options(joinedload(Objective.project)).all()
    key_results = KeyResult.query.filter(KeyResult.content.ilike(search_term)).options(joinedload(KeyResult.objective)).all()
    actions = ActionItem.query.filter(or_(ActionItem.content.ilike(search_term), ActionItem.report.ilike(search_term))).options(joinedload(ActionItem.key_result).joinedload(KeyResult.objective)).all()

    return render_template('search_results.html', page_name='search', query=query, tasks=tasks, objectives=objectives, key_results=key_results, actions=actions)


@bp.route('/uploads-manager')
@login_required
def uploads_manager():
    from sqlalchemy import asc, desc
    import shutil

    # --- Params & filter ---
    query_param   = request.args.get('q', type=str, default='')
    context_filter = request.args.get('context', 'all')
    user_filter    = request.args.get('user', 'all')
    page           = request.args.get('page', 1, type=int)
    per_page       = request.args.get('per_page', 20, type=int)  # cho phép chọn số file/trang

    files_q = UploadedFile.query
    if query_param:
        files_q = files_q.filter(UploadedFile.original_filename.ilike(f"%{query_param}%"))

    if context_filter == 'direct':
        files_q = files_q.filter(UploadedFile.upload_source == 'direct')
    elif context_filter == 'attachment':
        files_q = files_q.filter(UploadedFile.upload_source == 'attachment')

    if user_filter != 'all':
        files_q = files_q.filter(UploadedFile.uploader_id == user_filter)

    pagination = files_q.order_by(UploadedFile.upload_date.desc()).paginate(
        page=page, per_page=per_page, error_out=False
    )
    files = pagination.items

    # --- Thống kê & chart data ---
    # Dung lượng ổ dựa vào thư mục UPLOAD_FOLDER (ổ thật trên máy, tránh '/')
    try:
        # nên đo đúng ổ chứa thư mục upload
        disk_root = os.path.abspath(current_app.config['UPLOAD_FOLDER'])
        total, used, free = shutil.disk_usage(disk_root)
    except Exception:
        total = used = free = 0

    total_gb = total / (1024**3)
    used_gb  = used  / (1024**3)
    free_gb  = free  / (1024**3)

    # Tổng dung lượng các file đã upload
    all_files = UploadedFile.query.all()
    uploaded_total_bytes = sum((f.file_size or 0) for f in all_files)
    uploaded_gb = uploaded_total_bytes / (1024**3)
    other_gb = max(used_gb - uploaded_gb, 0.0)

    # Gom theo đuôi file
    from collections import defaultdict
    by_ext = defaultdict(lambda: {'count': 0, 'bytes': 0})
    for f in all_files:
        ext = (f.file_type or '').lower() or '(khác)'
        by_ext[ext]['count'] += 1
        by_ext[ext]['bytes'] += (f.file_size or 0)

    files_by_type = sorted(
        [(k, v['count'], v['bytes']) for k, v in by_ext.items()],
        key=lambda t: t[2],
        reverse=True
    )[:8]

    chart_data = {
        'total_gb': round(total_gb, 2),
        'uploaded_size_gb': round(uploaded_gb, 2),
        'other_data_size_gb': round(other_gb, 2),
        'free_space_gb': round(free_gb, 2),
        'files_by_type': files_by_type
    }

    # --- Context render ---
    all_users = User.query.all()
    context_choices = [
        {'value': 'all', 'text': 'All'},
        {'value': 'direct', 'text': 'Upload here'},
        {'value': 'attachment', 'text': 'Other tab'}
    ]

    return render_template(
        'uploads_manager.html',
        page_name='uploads_manager',
        files=files,
        pagination=pagination,
        query=query_param,
        context_filter=context_filter,
        user_filter=user_filter,
        context_choices=context_choices,
        all_users=all_users,
        chart_data=chart_data,
        total_files=len(all_files),
        per_page=per_page
    )


@bp.route('/projects')
@login_required
def projects_list():
    status_choices = ['Planned', 'Active', 'On Hold', 'Done']
    current_status = request.args.get('status', 'all')
    
    query = Project.query
    if current_status != 'all' and current_status in status_choices:
        query = query.filter(Project.status == current_status)
    
    projects = query.order_by(Project.name).all()

    return render_template('projects.html', projects=projects, page_name='projects', status_choices=status_choices, current_status=current_status)

@bp.route('/add-project', methods=['POST'])
@login_required
def add_project():
    name = request.form.get('name')
    if not name:
        flash('Tên dự án không được để trống.', 'danger')
    else:
        new_project = Project(name=name, description=request.form.get('description'), status=request.form.get('status', 'Active'))
        db.session.add(new_project)
        db.session.commit()
        flash('Created new project!', 'success')
    return redirect(url_for('main.projects_list'))

@bp.route('/update-project/<int:project_id>', methods=['POST'])
@login_required
def update_project(project_id):
    project = Project.query.get_or_404(project_id)
    new_status = request.form.get('status')
    if new_status:
        project.status = new_status
        db.session.commit()
        flash(f'Project status updated "{project.name}".', 'success')
    return redirect(url_for('main.projects_list', status=request.args.get('status', 'all')))

@bp.route('/delete-project/<int:project_id>', methods=['POST'])
@login_required
def delete_project(project_id):
    project = Project.query.get_or_404(project_id)
    try:
        db.session.delete(project)
        db.session.commit()
        flash(f'Project deleted "{project.name}".', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'Lỗi khi xóa dự án: {str(e)}', 'danger')
    return redirect(url_for('main.projects_list'))

# ==============================================================================
# BÁO CÁO & THỐNG KÊ
# ==============================================================================
@bp.route('/reports')
@login_required
def reports_page():
    selected_user_id = request.args.get('user_filter', 'all', type=str)
    selected_project_id = request.args.get('project_filter', 'all', type=str)
    selected_time_filter = request.args.get('time_filter', 'this_week', type=str)

    all_users = User.query.all()
    all_projects = Project.query.all()
    start_date, end_date = get_time_range_from_filter(selected_time_filter)
    base_task_query = Task.query.filter(Task.task_date.between(start_date, end_date))
    base_action_query = ActionItem.query.join(KeyResult).join(Objective).filter(Objective.start_date.between(start_date, end_date))

    if selected_project_id != 'all':
        base_action_query = base_action_query.filter(Objective.project_id == selected_project_id)
    
    users_to_report = all_users if selected_user_id == 'all' else [User.query.get(selected_user_id)]
    if not users_to_report[0]: users_to_report = []
    
    report_data = []
    for user in users_to_report:
        tasks_total = base_task_query.filter(Task.who_id == user.id).count()
        tasks_done = base_task_query.filter(Task.who_id == user.id, Task.status == 'Done').count()
        actions_total = base_action_query.filter(ActionItem.assignee_id == user.id).count()
        actions_done = base_action_query.filter(ActionItem.assignee_id == user.id, ActionItem.status == 'Done').count()
        report_data.append({
            'user': user,
            'tasks_total': tasks_total,
            'tasks_done': tasks_done,
            'task_completion_rate': (tasks_done / tasks_total * 100) if tasks_total > 0 else 0,
            'actions_total': actions_total,
            'actions_done': actions_done,
            'action_completion_rate': (actions_done / actions_total * 100) if actions_total > 0 else 0,
        })
    context = {
        'page_name': 'reports', 'users': all_users, 'projects': all_projects,
        'report_data': report_data, 'filters': {'user': selected_user_id, 'project': selected_project_id, 'time': selected_time_filter},
        'chart_labels': [d['user'].username for d in report_data],
        'chart_data_actions': [d['action_completion_rate'] for d in report_data],
        'chart_data_tasks': [d['task_completion_rate'] for d in report_data]
    }
    return render_template('reports.html', **context)

# ==============================================================================
# KANBAN BOARD
# ==============================================================================
# Trong file app/routes.py

# Đảm bảo bạn đã import thư viện này ở đầu file
from sqlalchemy import func

@bp.route('/kanban')
@login_required
def kanban_board():
    # --- BƯỚC 1: Lấy tham số 'overdue' từ URL ---
    period = request.args.get('period', 'week')
    user_filter = request.args.get('user_id', 'all')
    overdue_filter = request.args.get('overdue') # Sẽ là '1' hoặc None
    
    # === Xử lý các khoảng thời gian ===
    today = date.today()
    if period == 'day':
        start_date = end_date = today
    elif period == 'month':
        start_date = today.replace(day=1)
        end_date = start_date + timedelta(days=monthrange(today.year, today.month)[1] - 1)
    elif period == 'total':
        start_date = None
        end_date = None
    else: # Mặc định là 'week'
        start_date = today - timedelta(days=today.weekday())
        end_date = start_date + timedelta(days=6)

    # === Query tasks hiệu quả ===
    tasks_query = Task.query.options(joinedload(Task.assignee), joinedload(Task.attachments))
    
    if start_date and end_date:
        tasks_query = tasks_query.filter(Task.task_date.between(start_date, end_date))
        
    if user_filter != 'all':
        tasks_query = tasks_query.filter(Task.who_id == user_filter)
        
    # --- BƯỚC 2: Thêm điều kiện lọc 'overdue' vào query ---
    if overdue_filter == '1':
        tasks_query = tasks_query.filter(Task.task_date < today, Task.status != 'Done')
        
    all_tasks = tasks_query.order_by(Task.task_date.desc()).all()

    # === Chuẩn bị dữ liệu cho Giao diện ===
    status_meta = {
        'Pending': {'color': '#6c757d', 'icon': 'fa-solid fa-hourglass-half'},
        'In Progress': {'color': '#3B82F6', 'icon': 'fa-solid fa-person-digging'},
        'Review': {'color': '#ffc107', 'icon': 'fa-solid fa-magnifying-glass'},
        'Done': {'color': '#28a745', 'icon': 'fa-solid fa-circle-check'},
        'Drop': {'color': '#dc3545', 'icon': 'fa-solid fa-circle-xmark'}
    }
    kanban_statuses = list(status_meta.keys())

    tasks_by_status = defaultdict(list)
    for task in all_tasks:
        tasks_by_status[task.status].append(task)

    # === Chuẩn bị dữ liệu cho Biểu đồ ===
    summary_data = []
    tasks_by_user = defaultdict(lambda: {'tasks': [], 'status_counts': defaultdict(int)})
    for task in all_tasks:
        if task.assignee:
            user_key = task.assignee
            tasks_by_user[user_key]['tasks'].append(task)
            tasks_by_user[user_key]['status_counts'][task.status] += 1
            
    for user, data in tasks_by_user.items():
        summary_data.append({
            'user': user.to_dict(),
            'task_count': len(data['tasks']),
            'status_counts': dict(data['status_counts'])
        })
    summary_data.sort(key=lambda x: x['user']['username'])

    overall_status_counts = defaultdict(int)
    for task in all_tasks:
        overall_status_counts[task.status] += 1

    # --- BƯỚC 3: Trả trạng thái 'overdue' về cho template ---
    context = {
        'page_name': 'kanban',
        'title': 'Kanban Board',
        'grouped_tasks': dict(tasks_by_status),
        'kanban_statuses': kanban_statuses,
        'status_meta': status_meta,
        'users': User.query.order_by(User.username).all(),
        'today_date_obj': today,
        'filters': {
            'user': user_filter, 
            'period': period, 
            'overdue': overdue_filter # <--- THÊM DÒNG NÀY
        },
        'summary_data': summary_data,
        'overall_status_counts': dict(overall_status_counts)
    }
    
    return render_template('kanban.html', **context)

@bp.route('/update-task-status', methods=['POST'])
@login_required
def update_task_status():
    data = request.json
    try:
        task_id = data.get('taskId')
        new_status = data.get('newStatus')
        
        if not task_id or not new_status:
            return jsonify({'success': False, 'message': 'Miss infor taskId or newStatus'}), 400

        task = Task.query.get(task_id)
        if not task:
            return jsonify({'success': False, 'message': 'Cannot find job'}), 404
        
        old_status = task.status
        task.status = new_status
        
        # Ghi log lại hành động
        log_action = f"changed status '{task.what}' từ {old_status} sang {new_status}."
        db.session.add(Log(action=log_action, user_id=current_user.id))
        
        db.session.commit()
        return jsonify({'success': True, 'message': 'Status update success!'})
        
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'message': str(e)}), 500

@bp.route('/api/update-task-content/<int:task_id>', methods=['POST'])
@login_required
def update_task_content(task_id):
    task = Task.query.get_or_404(task_id)
    new_content = request.json.get('content', '').strip()
    if not new_content:
        return jsonify({'success': False, 'message': 'Must not empty'}), 400

    old_content, task.what = task.what, new_content
    db.session.add(Log(action=f"Update content job ID {task.id} from '{old_content}' to '{task.what}'.", user_id=current_user.id))
    db.session.commit()
    return jsonify({'success': True, 'message': 'Content updated!', 'new_content': task.what})

# ==============================================================================
# NOTES (KANBAN-STYLE)
# ==============================================================================
@bp.route('/notes')
@login_required
def notes():
    """Hiển thị trang ghi chú chính."""
    all_columns = Column.query.order_by(Column.position).all()
    if not all_columns:
        # Tạo các cột mặc định nếu chưa có
        default_titles = ['Idea', 'IP', 'Document','Seminar']
        for i, title in enumerate(default_titles):
            db.session.add(Column(name=title, position=i))
        db.session.commit()
        all_columns = Column.query.order_by(Column.position).all()

    notes_by_column_id = defaultdict(list)
    # Tải tất cả ghi chú và các file đính kèm liên quan
    all_notes = Note.query.options(joinedload(Note.creator),joinedload(Note.attachments)).all()
    print("\n--- DEBUGGING NOTES ---")
    for note in all_notes:
        # Lấy tên người dùng một cách an toàn
        creator_username = note.creator.username if note.creator else "N/A"
        
        print(f"Note ID: {note.id}, Title: '{note.title}', Creator ID: {note.creator_id}, Creator Username: {creator_username}")
    print("--- END DEBUGGING ---\n")
    # =======================================================

    for note in all_notes:
        notes_by_column_id[note.column_id].append(note)

    return render_template('notes.html', 
                           all_columns=all_columns, 
                           notes_by_column_id=notes_by_column_id, 
                           page_name='notes')

@bp.route('/api/notes', methods=['POST'])
@login_required
def create_note():
    """API tạo ghi chú mới, xử lý cả file đính kèm."""
    data = request.form
    title = data.get('title')
    column_id = data.get('column_id')

    if not title or not column_id:
        return jsonify({'success': False, 'message': 'Title and column must be fill'}), 400

    # Khai báo các thẻ và thuộc tính CSS được phép
    allowed_tags = list(bleach.ALLOWED_TAGS) + ['div', 'p', 'br', 'strong', 'em', 'u', 'ol', 'ul', 'li', 'img', 'a', 'span']
    allowed_attrs = {**bleach.ALLOWED_ATTRIBUTES, 'img': ['src', 'alt', 'style'], 'a': ['href', 'title'], 'span': ['style']}
    css_sanitizer = CSSSanitizer(allowed_css_properties=['color', 'background-color', 'text-align'])

    content = bleach.clean(
        data.get('content', ''), 
        tags=allowed_tags, 
        attributes=allowed_attrs,
        css_sanitizer=css_sanitizer # Thêm dòng này để hết cảnh báo
    )
    
    try:
        # BƯỚC SỬA LỖI QUAN TRỌNG NHẤT
        # 1. Khởi tạo Note mà không có creator_id
        note = Note(title=title, content=content, column_id=int(column_id))
        
        # 2. Gán creator_id một cách trực tiếp vào đối tượng
        note.creator_id = current_user.id 
        
        db.session.add(note)
        db.session.flush()

        files = request.files.getlist('attachments[]')
        for file in files:
            if file and file.filename != '':
                original_filename = secure_filename(file.filename)
                timestamp = datetime.now().strftime("%Y%m%d%H%M%S")
                saved_filename = f"{timestamp}_{original_filename}"
                file_path = os.path.join(current_app.config['UPLOAD_FOLDER'], saved_filename)
                file.save(file_path)
                file_size = os.path.getsize(file_path)
                _, file_ext = os.path.splitext(original_filename)
                new_file = UploadedFile(
                    original_filename=original_filename,
                    saved_filename=saved_filename,
                    file_type=file_ext.lower(),
                    file_size=file_size,
                    uploader_id=current_user.id,
                    upload_source='attachment',
                    note_id=note.id
                )
                db.session.add(new_file)

        db.session.commit()
        db.session.refresh(note)
        # Sửa lại message để không trả về object gây lỗi JSON
        return jsonify({'success': True, 'message': 'Create note sucess!'})
    except Exception as e:
        db.session.rollback()
        current_app.logger.error(f"Error creating note: {e}")
        return jsonify({'success': False, 'message': f'Error happend: {str(e)}'}), 500

@bp.route('/api/notes/<int:note_id>', methods=['PUT', 'POST'])
@login_required
def update_note(note_id):
    """
    API cập nhật ghi chú.
    Hỗ trợ cả JSON (cho việc kéo-thả đổi cột) và Form Data (cho việc sửa đầy đủ).
    """
    note = Note.query.get_or_404(note_id)

    try:
        if request.is_json:
            # Trường hợp kéo-thả thẻ ghi chú (chỉ cập nhật column_id)
            data = request.get_json()
            if 'column_id' in data:
                note.column_id = data['column_id']
        else:
            # Trường hợp sửa toàn bộ ghi chú từ form
            data = request.form
            
            allowed_tags = list(bleach.ALLOWED_TAGS) + ['div', 'p', 'br', 'strong', 'em', 'u', 'ol', 'ul', 'li', 'img', 'a', 'span']
            allowed_attrs = {**bleach.ALLOWED_ATTRIBUTES, 'img': ['src', 'alt', 'style'], 'a': ['href', 'title']}
            
            if 'title' in data:
                note.title = data['title']
            if 'content' in data:
                note.content = bleach.clean(data['content'], tags=allowed_tags, attributes=allowed_attrs)
            if 'column_id' in data:
                 note.column_id = data['column_id']

            # Xử lý file mới tải lên khi sửa
            files = request.files.getlist('attachments[]')
            for file in files:
                if file and file.filename != '':
                    original_filename = secure_filename(file.filename)
                    timestamp = datetime.now().strftime("%Y%m%d%H%M%S")
                    saved_filename = f"{timestamp}_{original_filename}"
                    file_path = os.path.join(current_app.config['UPLOAD_FOLDER'], saved_filename)
                    
                    file.save(file_path)
                    file_size = os.path.getsize(file_path)
                    _, file_ext = os.path.splitext(original_filename)

                    new_file = UploadedFile(
                        original_filename=original_filename,
                        saved_filename=saved_filename,
                        file_type=file_ext.lower(),
                        file_size=file_size,
                        uploader_id=current_user.id,
                        upload_source='attachment',
                        note_id=note.id
                    )
                    db.session.add(new_file)

        db.session.commit()
        db.session.refresh(note)
        return jsonify({'success': True, 'message': 'Updated note!'})
    except Exception as e:
        db.session.rollback()
        current_app.logger.error(f"Error updating note {note_id}: {e}")
        return jsonify({'success': False, 'message': f'Error happend: {str(e)}'}), 500

@bp.route('/api/notes/<int:note_id>', methods=['GET'])
@login_required
def get_note(note_id):
    """API lấy thông tin chi tiết của một ghi chú."""
    note = Note.query.options(joinedload(Note.attachments)).get_or_404(note_id)
    return jsonify({'success': True, 'note': note.to_dict()})

@bp.route('/api/notes/<int:note_id>', methods=['DELETE'])
@login_required
def delete_note(note_id):
    """API xóa một ghi chú."""
    note = Note.query.get_or_404(note_id)
    db.session.delete(note)
    db.session.commit()
    return jsonify({'success': True, 'message': 'Deleted note'})

@bp.route('/api/columns', methods=['POST'])
@login_required
def create_column():
    """API tạo một cột mới."""
    data = request.json
    if not data or not data.get('name'):
        return jsonify({'success': False, 'message': 'Column name must be not empty'}), 400
    max_pos = db.session.query(db.func.max(Column.position)).scalar() or -1
    new_column = Column(name=data['name'], position=max_pos + 1)
    db.session.add(new_column)
    db.session.commit()
    return jsonify({'success': True, 'column': {'id': new_column.id, 'name': new_column.name, 'position': new_column.position}})

@bp.route('/api/columns/<int:column_id>', methods=['DELETE'])
@login_required
def delete_column(column_id):
    """API xóa một cột và các ghi chú bên trong."""
    column = Column.query.get_or_404(column_id)
    db.session.delete(column)
    db.session.commit()
    return jsonify({'success': True, 'message': 'Column deleted'})

@bp.route('/api/columns/<int:column_id>/rename', methods=['POST'])
@login_required
def rename_column(column_id):
    """API đổi tên một cột."""
    column = Column.query.get_or_404(column_id)
    new_name = request.json.get('name', '').strip()
    if not new_name:
        return jsonify({'success': False, 'message': 'Column must be not empty'}), 400
    column.name = new_name
    db.session.commit()
    return jsonify({'success': True, 'new_name': column.name})

@bp.route('/api/columns/update-order', methods=['POST'])
@login_required
def update_column_order():
    """API cập nhật thứ tự các cột sau khi kéo-thả."""
    column_ids = request.json.get('order', [])
    for index, col_id in enumerate(column_ids):
        column = Column.query.get(col_id)
        if column:
            column.position = index
    db.session.commit()
    return jsonify({'success': True, 'message': 'Arranged column'})

@bp.route('/upload-image', methods=['POST'])
@login_required
def upload_image():
    """API để TinyMCE tải ảnh lên khi soạn thảo."""
    if 'file' in request.files:
        file = request.files['file']
        if file.filename != '':
            original_filename = secure_filename(file.filename)
            timestamp = datetime.now().strftime("%Y%m%d%H%M%S")
            saved_filename = f"{timestamp}_{original_filename}"
            file.save(os.path.join(current_app.config['UPLOAD_FOLDER'], saved_filename))
            return jsonify({'location': url_for('main.uploaded_file', filename=saved_filename)})
    return jsonify({'error': 'Upload failed'}), 400